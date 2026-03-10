"""Endpointy pro export prekladu — DOCX tabulka, CSV."""

import sys
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import csv
import io
import logging
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse

from services.project_store import get_project

logger = logging.getLogger(__name__)
router = APIRouter(tags=["export"])


@router.post("/api/projects/{project_id}/export/{format}")
async def api_export(project_id: str, format: str):
    """Export prekladu do CSV nebo JSON."""
    project = get_project(project_id)
    if not project:
        raise HTTPException(404, "Project not found")

    if format == "csv":
        return _export_csv(project)
    elif format == "json":
        return _export_json(project)
    elif format == "xlsx":
        return _export_xlsx(project)
    elif format == "xlsx-grouped":
        return _export_xlsx_grouped(project)
    else:
        raise HTTPException(400, f"Nepodporovany format: {format}. Podporovane: csv, json, xlsx, xlsx-grouped")


def _export_csv(project):
    """Export do CSV souboru."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["id", "original", "czech", "status", "category", "story_id"])

    for el in project.elements:
        writer.writerow([
            el.id,
            el.contents,
            el.czech or "",
            el.status.value if el.status else "",
            el.category.value if el.category else "",
            el.story_id or "",
        ])

    output.seek(0)
    filename = f"{project.id}_translations.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _export_xlsx(project):
    """Export do Excel XLSX."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Preklady"

    # Hlavicka
    headers = ["ID", "Vrstva", "Original", "Cesky preklad", "Status", "Kategorie"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    # Data
    for row, el in enumerate(project.elements, 2):
        ws.cell(row=row, column=1, value=el.id)
        ws.cell(row=row, column=2, value=el.layer_name or el.story_id or "")
        ws.cell(row=row, column=3, value=el.contents)
        ws.cell(row=row, column=4, value=el.czech or "")
        ws.cell(row=row, column=5, value=el.status.value if el.status else "")
        ws.cell(row=row, column=6, value=el.category.value if el.category else "")

    # Sirky sloupcu
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 12

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"{project.id}_translations.xlsx"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _export_xlsx_grouped(project):
    """Export do Excel XLSX — seskupeno podle kategorii."""
    from collections import defaultdict
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    wb.remove(wb.active)

    # Seskupit podle kategorie
    groups = defaultdict(list)
    uncategorized = []
    for el in project.elements:
        if el.category:
            groups[el.category.value].append(el)
        else:
            uncategorized.append(el)

    # Popis kategorii pro nazvy listu
    cat_labels = {
        "oceans_seas": "Oceany a more",
        "continents": "Kontinenty",
        "countries_full": "Staty",
        "countries_abbrev": "Staty (zkr.)",
        "regions": "Regiony",
        "cities": "Mesta",
        "water_bodies": "Vodni plochy",
        "landforms": "Tereny",
        "places": "Mista",
        "title": "Nadpisy",
        "info_boxes": "Info boxy",
        "labels": "Stitky",
        "annotations": "Anotace",
        "dates": "Data",
    }

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    cat_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    cat_font = Font(bold=True, size=11)

    # Jeden list se vsemi kategoriemi oddelenymi
    ws = wb.create_sheet("Podle kategorii")
    row = 1

    # Hlavicka
    headers = ["Original", "Cesky preklad", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
    row += 1

    # Serazene kategorie
    sorted_cats = sorted(groups.keys(), key=lambda c: cat_labels.get(c, c))
    for cat in sorted_cats:
        elements = groups[cat]
        # Radek s nazvem kategorie
        cell = ws.cell(row=row, column=1, value=cat_labels.get(cat, cat).upper())
        cell.fill = cat_fill
        cell.font = cat_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1

        for el in elements:
            ws.cell(row=row, column=1, value=el.contents)
            ws.cell(row=row, column=2, value=el.czech or "")
            ws.cell(row=row, column=3, value=el.status.value if el.status else "")
            row += 1

        row += 1  # prazdny radek mezi kategoriemi

    # Nekategorizovane na konec
    if uncategorized:
        cell = ws.cell(row=row, column=1, value="BEZ KATEGORIE")
        cell.fill = cat_fill
        cell.font = cat_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1
        for el in uncategorized:
            ws.cell(row=row, column=1, value=el.contents)
            ws.cell(row=row, column=2, value=el.czech or "")
            ws.cell(row=row, column=3, value=el.status.value if el.status else "")
            row += 1

    # Sirky
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 10

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"{project.id}_grouped.xlsx"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _export_json(project):
    """Export jako JSON pole prekladu."""
    data = []
    for el in project.elements:
        if el.czech:
            data.append({
                "id": el.id,
                "original": el.contents,
                "czech": el.czech,
                "status": el.status.value if el.status else None,
                "category": el.category.value if el.category else None,
            })
    return {"project_id": project.id, "total": len(data), "translations": data}
